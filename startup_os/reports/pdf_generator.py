# PDF & Export Generation Engine for StartupOS
# Uses WeasyPrint for complex investor-ready PDFs
# Falls back to Frappe's built-in jinja rendering

import os
import json
import frappe
from frappe.utils import today, flt, cstr, formatdate

TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "templates")


# ─────────────────────────────────────────────────────────────────
# 1. INVESTOR SUMMARY REPORT
# ─────────────────────────────────────────────────────────────────

def get_investor_report_data(company):
	"""Collect all data for the investor-ready PDF report."""

	# Runway
	runway = frappe.db.get_value(
		"Runway Model", {"company": company},
		["cash_balance", "base_burn", "runway_base", "runway_optimistic", "runway_pessimistic", "last_updated"],
		as_dict=True
	)

	# Unit Economics
	ue = frappe.db.get_value(
		"Unit Economics", {"company": company},
		["arpu", "cac", "ltv", "churn_rate", "payback_period", "magic_number"],
		as_dict=True
	)

	# Cap Table summary
	shareholders = frappe.get_all(
		"Shareholder",
		filters={"company": company},
		fields=["shareholder_name", "shareholder_type", "total_shares", "ownership_percent"],
		order_by="ownership_percent desc"
	)

	# Compliance Score
	compliance = frappe.db.get_value(
		"Compliance Score", {"company": company},
		["score", "last_computed"],
		as_dict=True
	)

	# Upcoming filings
	upcoming = frappe.get_all(
		"Regulatory Calendar",
		filters={"company": company, "status": "Upcoming"},
		fields=["filing_name", "form_type", "due_date"],
		order_by="due_date asc",
		limit=5
	)

	# Fundraising pipeline
	pipeline = frappe.get_all(
		"Fundraising Opportunity",
		filters={"company": company},
		fields=["stage", "amount", "probability"],
	)
	pipeline_summary = {
		"total_opportunities": len(pipeline),
		"total_pipeline_value": sum(flt(p.amount) for p in pipeline),
		"term_sheets": len([p for p in pipeline if p.stage == "Term Sheet"]),
		"closed_won": len([p for p in pipeline if p.stage == "Closed Won"]),
	}

	# ESOP summary
	esop_grants = frappe.get_all(
		"ESOP Grant",
		fields=["quantity"],
	)
	total_esop = sum(flt(g.quantity) for g in esop_grants)

	# DPIIT status
	dpiit = frappe.db.get_value(
		"DPIIT Application", {"company": company},
		["status", "application_number", "recognition_date"],
		as_dict=True
	)

	return frappe._dict({
		"company": company,
		"generated_on": formatdate(today(), "dd MMMM yyyy"),
		"runway": runway,
		"unit_economics": ue,
		"shareholders": shareholders,
		"compliance": compliance,
		"upcoming_filings": upcoming,
		"pipeline_summary": pipeline_summary,
		"total_esop_pool": total_esop,
		"dpiit": dpiit,
	})


@frappe.whitelist()
def generate_investor_report(company):
	"""Generate investor-ready PDF report. Returns base64-encoded PDF."""
	data = get_investor_report_data(company)
	html = _render_template("investor_report.html", data)
	return _html_to_pdf(html, f"Investor_Report_{company}_{today()}.pdf")


# ─────────────────────────────────────────────────────────────────
# 2. CAP TABLE PDF REPORT
# ─────────────────────────────────────────────────────────────────

@frappe.whitelist()
def generate_cap_table_report(company, as_of_date=None):
	"""Generate a formatted cap table PDF. Returns base64-encoded PDF."""
	as_of_date = as_of_date or today()

	# All shareholders
	shareholders = frappe.get_all(
		"Shareholder",
		filters={"company": company},
		fields=["name", "shareholder_name", "shareholder_type", "total_shares", "ownership_percent"],
		order_by="ownership_percent desc"
	)

	# Transactions per shareholder
	for sh in shareholders:
		transactions = frappe.get_all(
			"Equity Transaction",
			filters={"company": company, "shareholder": sh.name, "date": ["<=", as_of_date]},
			fields=["transaction_type", "share_type", "quantity", "price_per_share", "date"],
			order_by="date asc"
		)
		sh["transactions"] = transactions
		sh["share_classes"] = list(set(t.share_type for t in transactions))

	# ESOP pool
	esop_grants = frappe.get_all(
		"ESOP Grant",
		fields=["employee", "quantity", "grant_date"],
		order_by="grant_date desc"
	)
	total_esop_granted = sum(flt(g.quantity) for g in esop_grants)

	# Total shares count
	total_shares = sum(flt(sh.total_shares) for sh in shareholders)

	data = frappe._dict({
		"company": company,
		"as_of_date": formatdate(as_of_date, "dd MMMM yyyy"),
		"generated_on": formatdate(today(), "dd MMMM yyyy"),
		"shareholders": shareholders,
		"total_shares": total_shares,
		"esop_grants": esop_grants,
		"total_esop_granted": total_esop_granted,
	})

	html = _render_template("cap_table_report.html", data)
	return _html_to_pdf(html, f"Cap_Table_{company}_{today()}.pdf")


# ─────────────────────────────────────────────────────────────────
# 3. DPIIT COMPLIANCE REPORT
# ─────────────────────────────────────────────────────────────────

@frappe.whitelist()
def generate_dpiit_report(company):
	"""Generate DPIIT compliance report PDF."""

	dpiit = frappe.db.get_value(
		"DPIIT Application", {"company": company},
		["name", "application_number", "status", "date_of_application",
		 "certificate_number", "recognition_date"],
		as_dict=True
	)

	checklist = frappe.db.get_value(
		"Incorporation Checklist", {"company": company},
		["inc_20a", "pan", "tan", "gst", "shops_and_establishment", "status", "other_registrations"],
		as_dict=True
	)

	compliance_score = frappe.db.get_value(
		"Compliance Score", {"company": company},
		["score", "score_breakdown", "last_computed"],
		as_dict=True
	)

	all_filings = frappe.get_all(
		"Regulatory Calendar",
		filters={"company": company},
		fields=["filing_name", "form_type", "due_date", "status", "completion_date"],
		order_by="due_date asc"
	)

	filed = [f for f in all_filings if f.status == "Filed"]
	upcoming = [f for f in all_filings if f.status == "Upcoming"]
	overdue = [f for f in all_filings if f.status == "Overdue"]

	score_breakdown = []
	if compliance_score and compliance_score.score_breakdown:
		try:
			score_breakdown = json.loads(compliance_score.score_breakdown)
		except Exception:
			score_breakdown = []

	data = frappe._dict({
		"company": company,
		"generated_on": formatdate(today(), "dd MMMM yyyy"),
		"dpiit": dpiit,
		"checklist": checklist,
		"compliance_score": compliance_score,
		"score_breakdown": score_breakdown,
		"filed_filings": filed,
		"upcoming_filings": upcoming,
		"overdue_filings": overdue,
		"total_filings": len(all_filings),
	})

	html = _render_template("dpiit_report.html", data)
	return _html_to_pdf(html, f"DPIIT_Compliance_Report_{company}_{today()}.pdf")


# ─────────────────────────────────────────────────────────────────
# 4. AUDIT-READY EQUITY EXPORT (CSV / XLSX)
# ─────────────────────────────────────────────────────────────────

@frappe.whitelist()
def export_equity_audit(company, format="xlsx"):
	"""
	Export complete equity transaction log for audit purposes.
	Returns base64-encoded XLSX or CSV file.
	"""
	import base64, io

	# All transactions
	transactions = frappe.get_all(
		"Equity Transaction",
		filters={"company": company},
		fields=[
			"name", "date", "transaction_type", "share_type",
			"shareholder", "quantity", "price_per_share", "total_value"
		],
		order_by="date asc"
	)

	# Enrich with shareholder name
	for tx in transactions:
		tx["shareholder_name"] = frappe.db.get_value("Shareholder", tx.shareholder, "shareholder_name") or tx.shareholder

	# Compliance events
	compliance_events = frappe.get_all(
		"Regulatory Calendar",
		filters={"company": company},
		fields=["filing_name", "form_type", "due_date", "status", "completion_date"],
		order_by="due_date asc"
	)

	if format == "xlsx":
		try:
			import openpyxl
			from openpyxl.styles import Font, PatternFill, Alignment

			wb = openpyxl.Workbook()

			# Sheet 1: Equity Transactions
			ws1 = wb.active
			ws1.title = "Equity Transactions"

			headers = ["Transaction ID", "Date", "Type", "Share Type", "Shareholder", "Quantity", "Price/Share (₹)", "Total Value (₹)"]
			for col_idx, header in enumerate(headers, 1):
				cell = ws1.cell(row=1, column=col_idx, value=header)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

			for row_idx, tx in enumerate(transactions, 2):
				ws1.cell(row=row_idx, column=1, value=tx.name)
				ws1.cell(row=row_idx, column=2, value=cstr(tx.date))
				ws1.cell(row=row_idx, column=3, value=tx.transaction_type)
				ws1.cell(row=row_idx, column=4, value=tx.share_type)
				ws1.cell(row=row_idx, column=5, value=tx.shareholder_name)
				ws1.cell(row=row_idx, column=6, value=flt(tx.quantity))
				ws1.cell(row=row_idx, column=7, value=flt(tx.price_per_share))
				ws1.cell(row=row_idx, column=8, value=flt(tx.total_value))

			# Sheet 2: Compliance Events
			ws2 = wb.create_sheet("Compliance Events")
			headers2 = ["Filing Name", "Form", "Due Date", "Status", "Completion Date"]
			for col_idx, header in enumerate(headers2, 1):
				cell = ws2.cell(row=1, column=col_idx, value=header)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

			for row_idx, ev in enumerate(compliance_events, 2):
				ws2.cell(row=row_idx, column=1, value=ev.filing_name)
				ws2.cell(row=row_idx, column=2, value=ev.form_type)
				ws2.cell(row=row_idx, column=3, value=cstr(ev.due_date))
				ws2.cell(row=row_idx, column=4, value=ev.status)
				ws2.cell(row=row_idx, column=5, value=cstr(ev.completion_date) if ev.completion_date else "")

			# Sheet 3: Shareholder Summary
			ws3 = wb.create_sheet("Cap Table Summary")
			shareholders = frappe.get_all(
				"Shareholder",
				filters={"company": company},
				fields=["shareholder_name", "shareholder_type", "total_shares", "ownership_percent"],
				order_by="ownership_percent desc"
			)
			headers3 = ["Shareholder", "Type", "Total Shares", "Ownership %"]
			for col_idx, header in enumerate(headers3, 1):
				cell = ws3.cell(row=1, column=col_idx, value=header)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
			for row_idx, sh in enumerate(shareholders, 2):
				ws3.cell(row=row_idx, column=1, value=sh.shareholder_name)
				ws3.cell(row=row_idx, column=2, value=sh.shareholder_type)
				ws3.cell(row=row_idx, column=3, value=flt(sh.total_shares))
				ws3.cell(row=row_idx, column=4, value=flt(sh.ownership_percent))

			# Auto-width columns
			for ws in [ws1, ws2, ws3]:
				for col in ws.columns:
					max_len = max(len(str(cell.value or "")) for cell in col)
					ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

			buf = io.BytesIO()
			wb.save(buf)
			buf.seek(0)
			return {
				"filename": f"StartupOS_Audit_Export_{company}_{today()}.xlsx",
				"filecontent": base64.b64encode(buf.getvalue()).decode(),
				"type": "xlsx"
			}
		except ImportError:
			frappe.throw("openpyxl not installed. Run: pip install openpyxl")

	else:
		# CSV fallback
		import csv
		buf = io.StringIO()
		writer = csv.DictWriter(buf, fieldnames=["name", "date", "transaction_type", "share_type", "shareholder_name", "quantity", "price_per_share", "total_value"])
		writer.writeheader()
		writer.writerows([dict(tx) for tx in transactions])
		return {
			"filename": f"StartupOS_Equity_Audit_{company}_{today()}.csv",
			"filecontent": base64.b64encode(buf.getvalue().encode()).decode(),
			"type": "csv"
		}


# ─────────────────────────────────────────────────────────────────
# 5. FINANCIAL PROJECTIONS REPORT
# ─────────────────────────────────────────────────────────────────

@frappe.whitelist()
def generate_financial_projections(company, projection_months=24):
	"""Generate pitch-ready financial projections PDF."""

	runway = frappe.db.get_value(
		"Runway Model", {"company": company},
		["cash_balance", "base_burn", "runway_base", "runway_optimistic", "runway_pessimistic"],
		as_dict=True
	)

	ue = frappe.db.get_value(
		"Unit Economics", {"company": company},
		["arpu", "cac", "ltv", "churn_rate", "payback_period", "magic_number"],
		as_dict=True
	)

	# Build monthly projection table
	projections = []
	cash = flt(runway.cash_balance) if runway else 0
	monthly_burn = flt(runway.base_burn) if runway else 0
	arpu = flt(ue.arpu) if ue else 0
	churn = flt(ue.churn_rate) / 100 if ue else 0.02

	# Estimate starting customer base from current cash and burn rate
	# If arpu is set, seed with a conservative estimate (burn / arpu), else 0
	customers = int(monthly_burn / arpu) if (arpu and monthly_burn and arpu > 0) else 0

	from frappe.utils import add_months
	for month in range(1, int(projection_months) + 1):
		month_date = add_months(today(), month)
		# 8% MoM growth (conservative Seed-stage assumption)
		new_customers = int(customers * 1.08) + 1
		churned = int(customers * churn)
		customers = max(new_customers - churned, 0)
		mrr_proj = customers * arpu
		cash = cash - monthly_burn + mrr_proj

		projections.append({
			"month": month,
			"date": formatdate(month_date, "MMM yyyy"),
			"customers": customers,
			"mrr": mrr_proj,
			"cash_balance": cash,
			"burn": monthly_burn,
		})


	data = frappe._dict({
		"company": company,
		"generated_on": formatdate(today(), "dd MMMM yyyy"),
		"runway": runway,
		"unit_economics": ue,
		"projections": projections,
		"projection_months": projection_months,
	})

	html = _render_template("financial_projections.html", data)
	return _html_to_pdf(html, f"Financial_Projections_{company}_{today()}.pdf")


# ─────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────

def _render_template(template_name, data):
	"""Render a Jinja2 HTML template with the given data dict."""
	template_path = os.path.join(TEMPLATES_DIR, template_name)
	if not os.path.exists(template_path):
		frappe.throw(f"Report template not found: {template_name}")
	with open(template_path, "r") as f:
		template_str = f.read()
	from jinja2 import Environment
	env = Environment(autoescape=True)
	env.filters["inr"] = lambda v: f"₹{flt(v):,.0f}"
	env.filters["pct"] = lambda v: f"{flt(v):.2f}%"
	template = env.from_string(template_str)
	return template.render(**data)


def _html_to_pdf(html, filename):
	"""Convert HTML string to PDF and return base64-encoded content."""
	import base64
	try:
		from weasyprint import HTML
		pdf_bytes = HTML(string=html).write_pdf()
		return {
			"filename": filename,
			"filecontent": base64.b64encode(pdf_bytes).decode(),
			"type": "pdf"
		}
	except ImportError:
		# Fallback: return HTML content
		frappe.log_error("WeasyPrint not installed. Install via: pip install weasyprint", "PDF Generation")
		return {
			"filename": filename.replace(".pdf", ".html"),
			"filecontent": base64.b64encode(html.encode()).decode(),
			"type": "html"
		}
